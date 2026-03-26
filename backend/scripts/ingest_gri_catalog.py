"""
ingest_gri_catalog.py
=====================
Ingest GRI Standards taxonomy and ESRS-GRI data-point mapping into the
platform database.

Sources
-------
1. GRI Sustainability Taxonomy 2025 (XBRL)
   - SchemaElement → concrete data points (element_id, name, type)
   - Label Linkbase → human labels, documentation
   - Presentation Linkbase → disclosure hierarchy (standard ↔ disclosure ↔ DP)

2. EFRAG Draft ESRS-GRI Standards Data-Point Mapping
   - 12 sheets (one per ESRS standard)
   - Columns: ESRS, DR, Paragraph, Name, DataType, ... GRI Standard, Disclosure, Number, Name, Notes

Target tables
-------------
- gri_standards          (GRI disclosure catalog)
- gri_esrs_mapping       (row-level ESRS↔GRI mapping)
- csrd_esrs_catalog      (update gri_disclosure_ref quick-lookup column)
"""

import hashlib
import re
import uuid
import sys
import openpyxl
import psycopg2
import psycopg2.extras

DB_URL = "postgresql://postgres.kytzcbipsghprsqoalvi:KimiaAImpact2026@aws-1-us-east-2.pooler.supabase.com:5432/postgres"

GRI_TAXONOMY_PATH = r"C:\Users\SahilChopra\Downloads\gri-sustainability-taxonomy_2025-06-23.xlsx"
ESRS_GRI_MAPPING_PATH = r"C:\Users\SahilChopra\Downloads\draft-esrs-gri-standards-data-point-mapping.xlsx"


def make_uuid(seed: str) -> str:
    """Deterministic UUID5 from a seed string."""
    return str(uuid.uuid5(uuid.NAMESPACE_URL, seed))


# ── GRI Standard ↔ Topic mapping ───────────────────────────────────────
GRI_TOPIC_MAP = {
    "GRI 2": "General",
    "GRI 3": "General",
    "GRI 101": "Environmental",
    "GRI 201": "Economic",
    "GRI 202": "Economic",
    "GRI 203": "Economic",
    "GRI 204": "Economic",
    "GRI 205": "Governance",
    "GRI 206": "Governance",
    "GRI 207": "Economic",
    "GRI 301": "Environmental",
    "GRI 302": "Environmental",
    "GRI 303": "Environmental",
    "GRI 304": "Environmental",
    "GRI 305": "Environmental",
    "GRI 306": "Environmental",
    "GRI 308": "Environmental",
    "GRI 401": "Social",
    "GRI 402": "Social",
    "GRI 403": "Social",
    "GRI 404": "Social",
    "GRI 405": "Social",
    "GRI 406": "Social",
    "GRI 407": "Social",
    "GRI 408": "Social",
    "GRI 409": "Social",
    "GRI 410": "Social",
    "GRI 411": "Social",
    "GRI 413": "Social",
    "GRI 414": "Social",
    "GRI 415": "Governance",
    "GRI 416": "Social",
    "GRI 417": "Social",
    "GRI 418": "Social",
}


# ═══════════════════════════════════════════════════════════════════════
# PART 1 — Parse GRI Taxonomy → gri_standards table
# ═══════════════════════════════════════════════════════════════════════

def parse_gri_taxonomy(path: str) -> list[dict]:
    """
    Parse GRI Sustainability Taxonomy XBRL Excel to extract concrete
    disclosure data points with labels and documentation.
    """
    wb = openpyxl.load_workbook(path, read_only=True)

    # ── 1a. Schema elements ──
    schema = {}
    ws_schema = wb["SchemaElement"]
    for row in ws_schema.iter_rows(min_row=2, values_only=True):
        elem_name, elem_id, is_abstract, subst_group, period_type, dtype = (
            row[0], row[1], row[2], row[3], row[4], row[5]
        )
        if not elem_id:
            continue
        schema[str(elem_id).strip()] = {
            "element_id": str(elem_id).strip(),
            "element_name": str(elem_name).strip() if elem_name else "",
            "is_abstract": bool(is_abstract),
            "substitution_group": str(subst_group).strip() if subst_group else None,
            "period_type": str(period_type).strip() if period_type else None,
            "data_type": str(dtype).strip() if dtype else None,
        }

    # ── 1b. Labels ──
    ws_label = wb["Label Linkbase"]
    for row in ws_label.iter_rows(min_row=2, values_only=True):
        elem_id = str(row[0]).strip() if row[0] else None
        if not elem_id or elem_id not in schema:
            continue
        schema[elem_id]["label"] = str(row[2]).strip() if row[2] else None
        schema[elem_id]["verbose_label"] = str(row[8]).strip() if row[8] else None
        schema[elem_id]["documentation"] = str(row[6]).strip() if row[6] else None

    # ── 1c. Presentation linkbase → derive standard_code, disclosure_code, disclosure_name ──
    ws_pres = wb["Presentation Linkbase"]
    current_standard = None
    current_disclosure = None
    current_disclosure_name = None

    for row in ws_pres.iter_rows(min_row=2, values_only=True):
        elem_id = str(row[0]).strip() if row[0] else None
        label = str(row[1]).strip() if row[1] else ""

        # Section header like "[020000] GRI 2 - General Disclosures"
        if not elem_id and label.startswith("["):
            m = re.match(r"\[\d+\]\s+GRI\s+(\d+)\b", label)
            if m:
                current_standard = f"GRI {m.group(1)}"
            # Also detect "GRI 101" etc.
            m2 = re.match(r"\[\d+\]\s+GRI\s+(\d+)", label)
            if m2:
                current_standard = f"GRI {m2.group(1)}"
            continue

        # Disclosure header — verbose label like "2-1-a: Legal name"
        disc_match = re.match(r"Disclosure\s+(\d+-\d+)", label)
        if disc_match:
            current_disclosure = disc_match.group(1)
            current_disclosure_name = label
        elif label and re.match(r"\d+-\d+-[a-z]", label):
            # sub-item verbose label: "2-1-a: Legal name"
            dm = re.match(r"(\d+-\d+)", label)
            if dm:
                current_disclosure = dm.group(1)

        if elem_id and elem_id in schema:
            if current_standard:
                schema[elem_id]["standard_code"] = current_standard
            if current_disclosure:
                schema[elem_id]["disclosure_code"] = current_disclosure
            if current_disclosure_name:
                schema[elem_id]["disclosure_name"] = current_disclosure_name

    wb.close()

    # Build result list
    results = []
    for elem_id, rec in schema.items():
        std_code = rec.get("standard_code")
        topic = GRI_TOPIC_MAP.get(std_code, "General") if std_code else None
        results.append({
            "id": make_uuid(f"gri:{elem_id}"),
            "element_id": rec["element_id"],
            "element_name": rec["element_name"],
            "standard_code": std_code,
            "disclosure_code": rec.get("disclosure_code"),
            "disclosure_name": rec.get("disclosure_name"),
            "label": rec.get("label"),
            "verbose_label": rec.get("verbose_label"),
            "documentation": rec.get("documentation"),
            "data_type": rec["data_type"],
            "period_type": rec["period_type"],
            "is_abstract": rec["is_abstract"],
            "substitution_group": rec["substitution_group"],
            "topic_area": topic,
        })

    return results


# ═══════════════════════════════════════════════════════════════════════
# PART 2 — Parse ESRS-GRI mapping → gri_esrs_mapping + catalog update
# ═══════════════════════════════════════════════════════════════════════

# Column layout per sheet type:
# Topical (E1..G1) & ESRS 2:  col 0=ESRS, 1=DR, 2=Par, 3=RelAR, 4=Name, 5=DataType,
#                              6..9=appendix/voluntary,  10=GRI Standard, 11=GRI Disc, 12=GRI Number, 13=GRI Name, 14=Notes
# ESRS2 MDR:                  col 0=ESRS, 1=DR, 2=Par, 3=RelAR, 4=Name, 5=DataType,
#                              6=AppB, 7=AppC, 8=May[V],  9=GRI Standard, 10=GRI Disc, 11=GRI Number, 12=GRI Name, 13=Notes

SHEET_CONFIGS = {
    "ESRS 2":    {"gri_std_col": 10, "gri_disc_col": 11, "gri_num_col": 12, "gri_name_col": 13, "notes_col": 14, "esrs_code": "ESRS 2"},
    "ESRS2 MDR": {"gri_std_col": 9,  "gri_disc_col": 10, "gri_num_col": 11, "gri_name_col": 12, "notes_col": 13, "esrs_code": "ESRS 2"},
    "ESRS E1":   {"gri_std_col": 10, "gri_disc_col": 11, "gri_num_col": 12, "gri_name_col": 13, "notes_col": 14, "esrs_code": "E1"},
    "ESRS E2":   {"gri_std_col": 10, "gri_disc_col": 11, "gri_num_col": 12, "gri_name_col": 13, "notes_col": 14, "esrs_code": "E2"},
    "ESRS E3":   {"gri_std_col": 10, "gri_disc_col": 11, "gri_num_col": 12, "gri_name_col": 13, "notes_col": 14, "esrs_code": "E3"},
    "ESRS E4":   {"gri_std_col": 10, "gri_disc_col": 11, "gri_num_col": 12, "gri_name_col": 13, "notes_col": 14, "esrs_code": "E4"},
    "ESRS E5":   {"gri_std_col": 10, "gri_disc_col": 11, "gri_num_col": 12, "gri_name_col": 13, "notes_col": 14, "esrs_code": "E5"},
    "ESRS S1":   {"gri_std_col": 10, "gri_disc_col": 11, "gri_num_col": 12, "gri_name_col": 13, "notes_col": 14, "esrs_code": "S1"},
    "ESRS S2":   {"gri_std_col": 10, "gri_disc_col": 11, "gri_num_col": 12, "gri_name_col": 13, "notes_col": 14, "esrs_code": "S2"},
    "ESRS S3":   {"gri_std_col": 10, "gri_disc_col": 11, "gri_num_col": 12, "gri_name_col": 13, "notes_col": 14, "esrs_code": "S3"},
    "ESRS S4":   {"gri_std_col": 10, "gri_disc_col": 11, "gri_num_col": 12, "gri_name_col": 13, "notes_col": 14, "esrs_code": "S4"},
    "ESRS G1":   {"gri_std_col": 10, "gri_disc_col": 11, "gri_num_col": 12, "gri_name_col": 13, "notes_col": 14, "esrs_code": "G1"},
}


def _clean(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "\xa0", "None"):
        return None
    return s


def _make_esrs_indicator_code(esrs: str, dr: str, paragraph: str, row_idx: int) -> str:
    """Build indicator_code matching our catalog convention (e.g. E1-1_01)."""
    esrs_clean = _clean(esrs) or ""
    dr_clean = _clean(dr) or ""
    par_clean = _clean(paragraph) or ""

    # Normalize DR: " E1.GOV-3 " → "E1.GOV-3"
    dr_clean = dr_clean.strip()

    # Build a composite key for lookup against csrd_esrs_catalog
    # The catalog uses codes like E1-1_01, ESRS2.BP-1_01 etc.
    # We return the DR + paragraph for matching
    return f"{esrs_clean}|{dr_clean}|{par_clean}"


def parse_esrs_gri_mapping(path: str) -> list[dict]:
    """Parse the EFRAG draft ESRS-GRI mapping Excel."""
    wb = openpyxl.load_workbook(path, read_only=True)
    mappings = []
    row_counter = 0

    for sheet_name, cfg in SHEET_CONFIGS.items():
        if sheet_name not in wb.sheetnames:
            print(f"  WARN: sheet '{sheet_name}' not found, skipping")
            continue

        ws = wb[sheet_name]
        gri_std_col = cfg["gri_std_col"]
        gri_disc_col = cfg["gri_disc_col"]
        gri_num_col = cfg["gri_num_col"]
        gri_name_col = cfg["gri_name_col"]
        notes_col = cfg["notes_col"]
        esrs_code = cfg["esrs_code"]

        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or len(row) < 6:
                continue

            esrs_val = _clean(row[0])
            if not esrs_val or len(esrs_val) > 20:
                continue  # skip header/disclaimer rows

            dr_val = _clean(row[1])
            par_val = _clean(row[2])
            dp_name = _clean(row[4])

            # GRI columns
            gri_std = _clean(row[gri_std_col]) if gri_std_col < len(row) else None
            gri_disc = _clean(row[gri_disc_col]) if gri_disc_col < len(row) else None
            gri_num = _clean(row[gri_num_col]) if gri_num_col < len(row) else None
            gri_name = _clean(row[gri_name_col]) if gri_name_col < len(row) else None
            notes = _clean(row[notes_col]) if notes_col < len(row) else None

            # Skip rows with no GRI mapping
            if not gri_std:
                continue

            row_counter += 1

            # Handle "GRI 308; GRI 414" → split into multiple mappings
            gri_standards = [s.strip() for s in gri_std.split(";")]
            gri_disclosures = [gri_disc] * len(gri_standards) if len(gri_standards) > 1 else [gri_disc]

            for idx, gs in enumerate(gri_standards):
                gd = gri_disclosures[idx] if idx < len(gri_disclosures) else gri_disc
                seed = f"gri_map:{esrs_code}:{dr_val}:{par_val}:{gs}:{gd}:{gri_num}:{row_counter}"

                # Build composite GRI disclosure ref
                gri_ref = gs
                if gd:
                    gri_ref += f" {gd}"
                if gri_num:
                    gri_ref += f"-{gri_num}"

                mappings.append({
                    "id": make_uuid(seed),
                    "esrs_indicator_code": "",  # will be resolved later
                    "esrs_standard": esrs_code,
                    "esrs_dr": dr_val,
                    "esrs_paragraph": par_val,
                    "esrs_dp_name": dp_name,
                    "gri_standard": gs,
                    "gri_disclosure": gd,
                    "gri_sub_item": gri_num,
                    "gri_dp_name": gri_name,
                    "mapping_notes": notes,
                    "mapping_quality": "exact" if gri_disc else "thematic",
                    "_gri_ref": gri_ref,
                })

    wb.close()
    return mappings


# ═══════════════════════════════════════════════════════════════════════
# PART 3 — Resolve ESRS indicator codes + load into DB
# ═══════════════════════════════════════════════════════════════════════

def resolve_esrs_indicator_codes(cur, mappings: list[dict]):
    """
    For each mapping row, find the matching indicator_code in
    csrd_esrs_catalog by matching (standard_code, disclosure_requirement,
    paragraph_ref).
    """
    # Build lookup dict from catalog
    cur.execute("""
        SELECT indicator_code, standard_code, disclosure_requirement, paragraph_ref
        FROM csrd_esrs_catalog
    """)
    catalog = {}
    for ic, sc, dr, par in cur.fetchall():
        key = f"{sc}|{dr}|{par}"
        catalog[key] = ic
        # Also store without paragraph for DR-level matching
        dr_key = f"{sc}|{dr}"
        if dr_key not in catalog:
            catalog[dr_key] = ic

    resolved = 0
    for m in mappings:
        # Try exact match: standard|DR|paragraph
        key = f"{m['esrs_standard']}|{m['esrs_dr']}|{m['esrs_paragraph']}"
        if key in catalog:
            m["esrs_indicator_code"] = catalog[key]
            resolved += 1
        else:
            # Try DR-level match
            dr_key = f"{m['esrs_standard']}|{m['esrs_dr']}"
            if dr_key in catalog:
                m["esrs_indicator_code"] = catalog[dr_key]
                resolved += 1
            else:
                # Fallback: build a synthetic code
                dr_clean = (m["esrs_dr"] or "").strip().replace(" ", "")
                par_clean = (m["esrs_paragraph"] or "").strip().replace(" ", "_")
                m["esrs_indicator_code"] = f"{m['esrs_standard']}.{dr_clean}_{par_clean}"

    print(f"  Resolved {resolved}/{len(mappings)} mappings to catalog indicator codes")
    return mappings


def main():
    print("=" * 60)
    print("GRI Standards + ESRS-GRI Mapping Ingestion")
    print("=" * 60)

    conn = psycopg2.connect(DB_URL)
    conn.autocommit = True
    cur = conn.cursor()

    # ── Step 1: Ingest GRI Taxonomy ──────────────────────────────
    print("\n[1/4] Parsing GRI Sustainability Taxonomy 2025 ...")
    gri_records = parse_gri_taxonomy(GRI_TAXONOMY_PATH)
    print(f"  Parsed {len(gri_records)} elements ({sum(1 for r in gri_records if not r['is_abstract'])} concrete)")

    # Clear and reload
    cur.execute("DELETE FROM gri_standards")
    cols = [
        "id", "element_id", "element_name", "standard_code", "disclosure_code",
        "disclosure_name", "label", "verbose_label", "documentation", "data_type",
        "period_type", "is_abstract", "substitution_group", "topic_area",
    ]
    template = "(" + ",".join(["%s"] * len(cols)) + ")"
    values = [
        tuple(r.get(c) for c in cols)
        for r in gri_records
    ]

    batch_size = 200
    for i in range(0, len(values), batch_size):
        batch = values[i:i + batch_size]
        psycopg2.extras.execute_values(
            cur,
            f"INSERT INTO gri_standards ({','.join(cols)}) VALUES %s ON CONFLICT (id) DO NOTHING",
            batch,
            template=template,
        )
    cur.execute("SELECT COUNT(*) FROM gri_standards")
    print(f"  Loaded {cur.fetchone()[0]} rows into gri_standards")

    # ── Step 2: Ingest ESRS-GRI Mapping ──────────────────────────
    print("\n[2/4] Parsing ESRS-GRI data-point mapping ...")
    mapping_records = parse_esrs_gri_mapping(ESRS_GRI_MAPPING_PATH)
    print(f"  Parsed {len(mapping_records)} mapping rows")

    # ── Step 3: Resolve indicator codes ──────────────────────────
    print("\n[3/4] Resolving ESRS indicator codes ...")
    mapping_records = resolve_esrs_indicator_codes(cur, mapping_records)

    # Insert mapping rows
    cur.execute("DELETE FROM gri_esrs_mapping")
    map_cols = [
        "id", "esrs_indicator_code", "esrs_standard", "esrs_dr", "esrs_paragraph",
        "esrs_dp_name", "gri_standard", "gri_disclosure", "gri_sub_item",
        "gri_dp_name", "mapping_notes", "mapping_quality",
    ]
    map_template = "(" + ",".join(["%s"] * len(map_cols)) + ")"
    map_values = [
        tuple(r.get(c) for c in map_cols)
        for r in mapping_records
    ]

    for i in range(0, len(map_values), batch_size):
        batch = map_values[i:i + batch_size]
        psycopg2.extras.execute_values(
            cur,
            f"INSERT INTO gri_esrs_mapping ({','.join(map_cols)}) VALUES %s ON CONFLICT (id) DO NOTHING",
            batch,
            template=map_template,
        )
    cur.execute("SELECT COUNT(*) FROM gri_esrs_mapping")
    print(f"  Loaded {cur.fetchone()[0]} rows into gri_esrs_mapping")

    # ── Step 4: Update gri_disclosure_ref on csrd_esrs_catalog ───
    print("\n[4/4] Updating csrd_esrs_catalog.gri_disclosure_ref ...")

    # For each catalog row, pick the first GRI mapping as the primary ref
    cur.execute("""
        UPDATE csrd_esrs_catalog c
        SET gri_disclosure_ref = sub.gri_ref
        FROM (
            SELECT DISTINCT ON (esrs_indicator_code)
                esrs_indicator_code,
                CASE
                    WHEN gri_disclosure IS NOT NULL AND gri_sub_item IS NOT NULL
                    THEN gri_standard || ' ' || gri_disclosure || '-' || gri_sub_item
                    WHEN gri_disclosure IS NOT NULL
                    THEN gri_standard || ' ' || gri_disclosure
                    ELSE gri_standard
                END as gri_ref
            FROM gri_esrs_mapping
            ORDER BY esrs_indicator_code, gri_standard, gri_disclosure
        ) sub
        WHERE c.indicator_code = sub.esrs_indicator_code
    """)
    updated = cur.rowcount
    print(f"  Updated {updated} catalog rows with gri_disclosure_ref")

    # ── Summary ──────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)

    cur.execute("SELECT COUNT(*) FROM gri_standards WHERE is_abstract = false")
    concrete = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM gri_standards")
    total_gri = cur.fetchone()[0]
    print(f"  gri_standards:     {total_gri} total ({concrete} concrete data points)")

    cur.execute("SELECT COUNT(*) FROM gri_esrs_mapping")
    total_map = cur.fetchone()[0]
    print(f"  gri_esrs_mapping:  {total_map} mapping rows")

    cur.execute("SELECT COUNT(*) FROM csrd_esrs_catalog WHERE gri_disclosure_ref IS NOT NULL")
    gri_linked = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM csrd_esrs_catalog")
    total_cat = cur.fetchone()[0]
    print(f"  ESRS catalog GRI:  {gri_linked}/{total_cat} ({gri_linked*100//max(total_cat,1)}%) linked to GRI")

    # Breakdown by standard
    print("\n  GRI mapping by ESRS standard:")
    cur.execute("""
        SELECT esrs_standard, COUNT(*) as cnt,
               COUNT(DISTINCT gri_standard) as gri_stds,
               COUNT(DISTINCT gri_disclosure) as gri_discs
        FROM gri_esrs_mapping
        GROUP BY esrs_standard
        ORDER BY esrs_standard
    """)
    for row in cur.fetchall():
        print(f"    {row[0]:8s}  {row[1]:4d} mappings → {row[2]:2d} GRI standards, {row[3]:2d} disclosures")

    # GRI standard coverage
    print("\n  GRI standards referenced:")
    cur.execute("""
        SELECT gri_standard, COUNT(*) as cnt
        FROM gri_esrs_mapping
        GROUP BY gri_standard
        ORDER BY COUNT(*) DESC
    """)
    for row in cur.fetchall():
        print(f"    {row[0]:20s}  {row[1]:4d} mappings")

    cur.close()
    conn.close()
    print("\nDone.")


if __name__ == "__main__":
    main()
