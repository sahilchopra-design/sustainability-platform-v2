"""
Universal Export Service for Climate & Real Estate Risk Platform
Exports analysis results to PDF and Excel formats for all modules
"""

import io
import json
from datetime import datetime, timezone
from typing import Dict, List, Any, Optional
from decimal import Decimal
from uuid import UUID

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, 
    PageBreak, Image, HRFlowable
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Excel generation
import xlsxwriter
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import BarChart, Reference, PieChart


class DecimalEncoder(json.JSONEncoder):
    """JSON encoder that handles Decimal and UUID types."""
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        if isinstance(obj, UUID):
            return str(obj)
        if isinstance(obj, datetime):
            return obj.isoformat()
        return super().default(obj)


class ExportService:
    """Universal export service for generating PDF and Excel reports."""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Setup custom paragraph styles."""
        self.styles.add(ParagraphStyle(
            name='ReportTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#1e3a5f')
        ))
        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceBefore=20,
            spaceAfter=10,
            textColor=colors.HexColor('#2563eb'),
            borderPadding=(0, 0, 5, 0),
        ))
        self.styles.add(ParagraphStyle(
            name='SubHeader',
            parent=self.styles['Heading3'],
            fontSize=11,
            spaceBefore=10,
            spaceAfter=5,
            textColor=colors.HexColor('#475569'),
        ))
        self.styles.add(ParagraphStyle(
            name='ReportBodyText',
            parent=self.styles['Normal'],
            fontSize=10,
            spaceAfter=6,
        ))
        self.styles.add(ParagraphStyle(
            name='MetricLabel',
            parent=self.styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#64748b'),
        ))
        self.styles.add(ParagraphStyle(
            name='MetricValue',
            parent=self.styles['Normal'],
            fontSize=12,
            fontName='Helvetica-Bold',
            textColor=colors.HexColor('#0f172a'),
        ))
    
    # ==================== PDF Generation ====================
    
    def generate_portfolio_analytics_pdf(self, data: Dict) -> io.BytesIO:
        """Generate PDF report for Portfolio Analytics."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            rightMargin=40, leftMargin=40,
            topMargin=40, bottomMargin=40
        )
        
        story = []
        
        # Title
        story.append(Paragraph("Portfolio Analytics Report", self.styles['ReportTitle']))
        story.append(Paragraph(
            f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
            self.styles['MetricLabel']
        ))
        story.append(Spacer(1, 20))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e2e8f0')))
        story.append(Spacer(1, 20))
        
        # Executive Summary
        if 'executive_summary' in data:
            story.append(Paragraph("Executive Summary", self.styles['SectionHeader']))
            summary = data['executive_summary']
            summary_data = [
                ['Portfolio', summary.get('portfolio_name', 'N/A')],
                ['Total Value', f"${summary.get('total_value', 0):,.0f}"],
                ['Properties', str(summary.get('property_count', 0))],
                ['Average Risk Score', f"{summary.get('avg_risk_score', 0):.1f}"],
                ['Stranded Assets', str(summary.get('stranded_assets', 0))],
            ]
            table = Table(summary_data, colWidths=[2*inch, 3*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8fafc')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1e293b')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ]))
            story.append(table)
            story.append(Spacer(1, 20))
        
        # Key Findings
        if 'key_findings' in data.get('executive_summary', {}):
            story.append(Paragraph("Key Findings", self.styles['SubHeader']))
            for finding in data['executive_summary']['key_findings']:
                story.append(Paragraph(f"• {finding}", self.styles['ReportBodyText']))
            story.append(Spacer(1, 15))
        
        # Portfolio Overview
        if 'portfolio_overview' in data:
            story.append(Paragraph("Portfolio Overview", self.styles['SectionHeader']))
            overview = data['portfolio_overview']
            overview_data = [
                ['Portfolio Type', overview.get('portfolio_type', 'N/A')],
                ['Investment Strategy', overview.get('investment_strategy', 'N/A')],
                ['AUM', f"${overview.get('aum', 0):,.0f}"],
                ['Currency', overview.get('currency', 'USD')],
                ['Total Income', f"${overview.get('total_income', 0):,.0f}"],
                ['Yield', f"{overview.get('yield', 0):.2f}%"],
            ]
            table = Table(overview_data, colWidths=[2*inch, 3*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8fafc')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(table)
            story.append(Spacer(1, 20))
        
        # Property Details
        if 'property_details' in data:
            story.append(PageBreak())
            story.append(Paragraph("Property Details", self.styles['SectionHeader']))
            
            headers = ['Property', 'Type', 'Location', 'Value ($M)', 'Risk']
            prop_data = [headers]
            
            for prop in data['property_details'][:20]:  # Limit to 20 properties
                prop_data.append([
                    prop.get('name', 'N/A')[:25],
                    prop.get('type', 'N/A'),
                    prop.get('location', 'N/A')[:20],
                    f"{prop.get('value', 0)/1e6:.1f}",
                    str(prop.get('risk_score', 'N/A')),
                ])
            
            table = Table(prop_data, colWidths=[2*inch, 1*inch, 1.5*inch, 1*inch, 0.7*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (3, 1), (4, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
            ]))
            story.append(table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def generate_sustainability_pdf(self, data: Dict) -> io.BytesIO:
        """Generate PDF report for Sustainability Assessment."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            rightMargin=40, leftMargin=40,
            topMargin=40, bottomMargin=40
        )
        
        story = []
        
        # Title
        story.append(Paragraph("Sustainability Assessment Report", self.styles['ReportTitle']))
        story.append(Paragraph(
            f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
            self.styles['MetricLabel']
        ))
        story.append(Spacer(1, 20))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#10b981')))
        story.append(Spacer(1, 20))
        
        # Assessment Type Header
        assessment_type = data.get('assessment_type', 'Sustainability').upper()
        story.append(Paragraph(f"{assessment_type} Assessment Results", self.styles['SectionHeader']))
        
        # Score Summary
        if 'total_score' in data or 'weighted_score' in data:
            score = data.get('total_score') or data.get('weighted_score', 0)
            rating = data.get('star_rating') or data.get('rating') or data.get('certification_level', 'N/A')
            
            score_data = [
                ['Overall Score', f"{score:.1f}" if isinstance(score, (int, float, Decimal)) else str(score)],
                ['Rating/Level', str(rating).replace('_', ' ').title()],
            ]
            
            if 'percentile_rank' in data:
                score_data.append(['Percentile Rank', f"{data['percentile_rank']}%"])
            
            table = Table(score_data, colWidths=[2*inch, 2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0fdf4')),
                ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#16a34a')),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 12),
                ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#86efac')),
            ]))
            story.append(table)
            story.append(Spacer(1, 20))
        
        # Value Impact
        story.append(Paragraph("Value Impact Analysis", self.styles['SectionHeader']))
        value_data = [
            ['Metric', 'Value'],
            ['Rent Premium', f"+{data.get('estimated_rent_premium_percent', 0):.1f}%"],
            ['Value Premium', f"+{data.get('estimated_value_premium_percent', 0):.1f}%"],
        ]
        if data.get('estimated_value_impact'):
            value_data.append(['Est. Value Impact', f"${data['estimated_value_impact']:,.0f}"])
        if data.get('cap_rate_compression_bps'):
            value_data.append(['Cap Rate Compression', f"{data['cap_rate_compression_bps']} bps"])
        
        table = Table(value_data, colWidths=[2.5*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(table)
        story.append(Spacer(1, 20))
        
        # Recommendations
        if 'improvement_recommendations' in data or 'improvement_priorities' in data:
            story.append(Paragraph("Recommendations", self.styles['SectionHeader']))
            recommendations = data.get('improvement_recommendations', []) or \
                             [p.get('category', '') for p in data.get('improvement_priorities', [])]
            for rec in recommendations[:5]:
                story.append(Paragraph(f"• {rec}", self.styles['ReportBodyText']))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def generate_stranded_assets_pdf(self, data: Dict) -> io.BytesIO:
        """Generate PDF report for Stranded Asset Analysis."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            rightMargin=40, leftMargin=40,
            topMargin=40, bottomMargin=40
        )
        
        story = []
        
        # Title
        story.append(Paragraph("Stranded Asset Analysis Report", self.styles['ReportTitle']))
        story.append(Paragraph(
            f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
            self.styles['MetricLabel']
        ))
        story.append(Spacer(1, 20))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#ef4444')))
        story.append(Spacer(1, 20))
        
        # Risk Summary
        story.append(Paragraph("Risk Summary", self.styles['SectionHeader']))
        
        risk_data = [
            ['Metric', 'Value'],
            ['Risk Score', f"{data.get('stranding_risk_score', 0):.2f}"],
            ['Risk Category', data.get('risk_category', 'N/A').upper()],
            ['Stranded Volume', f"{data.get('stranded_volume_pct', 0):.1f}%"],
            ['NPV Impact', f"${data.get('npv_impact', 0):,.0f}"],
        ]
        
        table = Table(risk_data, colWidths=[2.5*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc2626')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#fecaca')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(table)
        story.append(Spacer(1, 20))
        
        # Key Drivers
        if 'key_drivers' in data:
            story.append(Paragraph("Key Risk Drivers", self.styles['SectionHeader']))
            for driver in data['key_drivers']:
                story.append(Paragraph(f"• {driver}", self.styles['ReportBodyText']))
            story.append(Spacer(1, 15))
        
        # Recommendations
        if 'recommendations' in data:
            story.append(Paragraph("Recommendations", self.styles['SectionHeader']))
            for rec in data['recommendations']:
                story.append(Paragraph(f"• {rec}", self.styles['ReportBodyText']))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def generate_scenario_analysis_pdf(self, data: Dict) -> io.BytesIO:
        """Generate PDF report for Scenario Analysis."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            rightMargin=40, leftMargin=40,
            topMargin=40, bottomMargin=40
        )
        
        story = []
        
        # Title
        story.append(Paragraph("Scenario Analysis Report", self.styles['ReportTitle']))
        story.append(Paragraph(
            f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
            self.styles['MetricLabel']
        ))
        story.append(Spacer(1, 20))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#8b5cf6')))
        story.append(Spacer(1, 20))
        
        # Scenario Comparison
        if 'comparison_table' in data:
            story.append(Paragraph("Scenario Comparison", self.styles['SectionHeader']))
            
            headers = ['Scenario', 'Total Value', 'Change %', 'Risk Score', 'VaR 95%']
            table_data = [headers]
            
            for row in data['comparison_table']:
                table_data.append([
                    row.get('scenario_name', 'N/A'),
                    f"${row.get('total_value', 0)/1e6:.1f}M",
                    f"{row.get('value_change_pct', 0):+.1f}%",
                    f"{row.get('avg_risk_score', 0):.1f}",
                    f"${row.get('var_95', 0)/1e6:.1f}M",
                ])
            
            table = Table(table_data, colWidths=[1.8*inch, 1.2*inch, 1*inch, 0.9*inch, 1.2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7c3aed')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
            ]))
            story.append(table)
            story.append(Spacer(1, 20))
        
        # Best/Worst Summary
        if 'best_scenario' in data or 'worst_scenario' in data:
            story.append(Paragraph("Analysis Summary", self.styles['SectionHeader']))
            summary_data = []
            if 'best_scenario' in data:
                summary_data.append(['Best Scenario', data['best_scenario']])
            if 'worst_scenario' in data:
                summary_data.append(['Worst Scenario', data['worst_scenario']])
            if 'value_spread' in data:
                summary_data.append(['Value Spread', f"${data['value_spread']/1e6:.1f}M"])
            
            table = Table(summary_data, colWidths=[2*inch, 3*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#faf5ff')),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e9d5ff')),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(table)
            story.append(Spacer(1, 15))
        
        # Key Insights
        if 'key_insights' in data:
            story.append(Paragraph("Key Insights", self.styles['SectionHeader']))
            for insight in data['key_insights']:
                story.append(Paragraph(f"• {insight}", self.styles['ReportBodyText']))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    # ==================== Excel Generation ====================
    
    def generate_portfolio_analytics_excel(self, data: Dict) -> io.BytesIO:
        """Generate Excel report for Portfolio Analytics."""
        buffer = io.BytesIO()
        workbook = xlsxwriter.Workbook(buffer, {'in_memory': True})
        
        # Define formats
        header_format = workbook.add_format({
            'bold': True, 'bg_color': '#2563eb', 'font_color': 'white',
            'align': 'center', 'valign': 'vcenter', 'border': 1
        })
        label_format = workbook.add_format({
            'bold': True, 'bg_color': '#f8fafc', 'border': 1
        })
        value_format = workbook.add_format({'border': 1, 'num_format': '#,##0'})
        currency_format = workbook.add_format({'border': 1, 'num_format': '$#,##0'})
        title_format = workbook.add_format({
            'bold': True, 'font_size': 16, 'font_color': '#1e3a5f'
        })
        
        # Summary Sheet
        summary_sheet = workbook.add_worksheet('Summary')
        summary_sheet.set_column('A:A', 25)
        summary_sheet.set_column('B:B', 30)
        
        summary_sheet.write('A1', 'Portfolio Analytics Report', title_format)
        summary_sheet.write('A2', f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}")
        
        row = 4
        if 'executive_summary' in data:
            summary = data['executive_summary']
            summary_sheet.write(row, 0, 'Portfolio Name', label_format)
            summary_sheet.write(row, 1, summary.get('portfolio_name', 'N/A'))
            row += 1
            summary_sheet.write(row, 0, 'Total Value', label_format)
            summary_sheet.write(row, 1, summary.get('total_value', 0), currency_format)
            row += 1
            summary_sheet.write(row, 0, 'Property Count', label_format)
            summary_sheet.write(row, 1, summary.get('property_count', 0), value_format)
            row += 1
            summary_sheet.write(row, 0, 'Avg Risk Score', label_format)
            summary_sheet.write(row, 1, summary.get('avg_risk_score', 0))
            row += 1
            summary_sheet.write(row, 0, 'Stranded Assets', label_format)
            summary_sheet.write(row, 1, summary.get('stranded_assets', 0), value_format)
        
        # Properties Sheet
        if 'property_details' in data:
            prop_sheet = workbook.add_worksheet('Properties')
            prop_sheet.set_column('A:A', 30)
            prop_sheet.set_column('B:F', 15)
            
            headers = ['Property Name', 'Type', 'Location', 'Value', 'Risk Score', 'Certifications']
            for col, header in enumerate(headers):
                prop_sheet.write(0, col, header, header_format)
            
            for row_idx, prop in enumerate(data['property_details'], start=1):
                prop_sheet.write(row_idx, 0, prop.get('name', 'N/A'))
                prop_sheet.write(row_idx, 1, prop.get('type', 'N/A'))
                prop_sheet.write(row_idx, 2, prop.get('location', 'N/A'))
                prop_sheet.write(row_idx, 3, prop.get('value', 0), currency_format)
                prop_sheet.write(row_idx, 4, prop.get('risk_score', 'N/A'))
                certs = prop.get('certifications', [])
                prop_sheet.write(row_idx, 5, ', '.join(certs) if certs else 'None')
        
        workbook.close()
        buffer.seek(0)
        return buffer
    
    def generate_sustainability_excel(self, data: Dict) -> io.BytesIO:
        """Generate Excel report for Sustainability Assessment."""
        buffer = io.BytesIO()
        workbook = xlsxwriter.Workbook(buffer, {'in_memory': True})
        
        # Formats
        header_format = workbook.add_format({
            'bold': True, 'bg_color': '#16a34a', 'font_color': 'white',
            'align': 'center', 'valign': 'vcenter', 'border': 1
        })
        label_format = workbook.add_format({
            'bold': True, 'bg_color': '#f0fdf4', 'border': 1
        })
        title_format = workbook.add_format({
            'bold': True, 'font_size': 16, 'font_color': '#16a34a'
        })
        
        # Unused but defined for potential future use
        _ = header_format  # Suppress linter warning
        _ = label_format
        
        # Assessment Results Sheet
        sheet = workbook.add_worksheet('Assessment')
        sheet.set_column('A:A', 25)
        sheet.set_column('B:B', 25)
        
        assessment_type = data.get('assessment_type', 'Sustainability').upper()
        sheet.write('A1', f'{assessment_type} Assessment Report', title_format)
        sheet.write('A2', f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}")
        
        row = 4
        
        # Scores
        sheet.write(row, 0, 'Score Summary', header_format)
        sheet.write(row, 1, '', header_format)
        row += 1
        
        score = data.get('total_score') or data.get('weighted_score', 0)
        rating = data.get('star_rating') or data.get('rating') or data.get('certification_level', 'N/A')
        
        sheet.write(row, 0, 'Overall Score', label_format)
        sheet.write(row, 1, float(score) if isinstance(score, (Decimal,)) else score)
        row += 1
        
        sheet.write(row, 0, 'Rating/Level', label_format)
        sheet.write(row, 1, str(rating).replace('_', ' ').title())
        row += 1
        
        if 'percentile_rank' in data:
            sheet.write(row, 0, 'Percentile Rank', label_format)
            sheet.write(row, 1, f"{data['percentile_rank']}%")
            row += 1
        
        row += 1
        
        # Value Impact
        sheet.write(row, 0, 'Value Impact', header_format)
        sheet.write(row, 1, '', header_format)
        row += 1
        
        sheet.write(row, 0, 'Rent Premium', label_format)
        sheet.write(row, 1, f"+{data.get('estimated_rent_premium_percent', 0):.1f}%")
        row += 1
        
        sheet.write(row, 0, 'Value Premium', label_format)
        sheet.write(row, 1, f"+{data.get('estimated_value_premium_percent', 0):.1f}%")
        row += 1
        
        if data.get('estimated_value_impact'):
            sheet.write(row, 0, 'Est. Value Impact', label_format)
            sheet.write(row, 1, f"${data['estimated_value_impact']:,.0f}")
            row += 1
        
        if data.get('cap_rate_compression_bps'):
            sheet.write(row, 0, 'Cap Rate Compression', label_format)
            sheet.write(row, 1, f"{data['cap_rate_compression_bps']} bps")
        
        workbook.close()
        buffer.seek(0)
        return buffer
    
    def generate_stranded_assets_excel(self, data: Dict) -> io.BytesIO:
        """Generate Excel report for Stranded Asset Analysis."""
        buffer = io.BytesIO()
        workbook = xlsxwriter.Workbook(buffer, {'in_memory': True})
        
        # Formats
        header_format = workbook.add_format({
            'bold': True, 'bg_color': '#dc2626', 'font_color': 'white',
            'align': 'center', 'border': 1
        })
        label_format = workbook.add_format({
            'bold': True, 'bg_color': '#fef2f2', 'border': 1
        })
        title_format = workbook.add_format({
            'bold': True, 'font_size': 16, 'font_color': '#dc2626'
        })
        
        sheet = workbook.add_worksheet('Stranded Asset Analysis')
        sheet.set_column('A:A', 25)
        sheet.set_column('B:B', 25)
        
        sheet.write('A1', 'Stranded Asset Analysis Report', title_format)
        sheet.write('A2', f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}")
        
        row = 4
        sheet.write(row, 0, 'Risk Summary', header_format)
        sheet.write(row, 1, '', header_format)
        row += 1
        
        sheet.write(row, 0, 'Risk Score', label_format)
        sheet.write(row, 1, f"{data.get('stranding_risk_score', 0):.2f}")
        row += 1
        
        sheet.write(row, 0, 'Risk Category', label_format)
        sheet.write(row, 1, data.get('risk_category', 'N/A').upper())
        row += 1
        
        sheet.write(row, 0, 'Stranded Volume', label_format)
        sheet.write(row, 1, f"{data.get('stranded_volume_pct', 0):.1f}%")
        row += 1
        
        sheet.write(row, 0, 'NPV Impact', label_format)
        sheet.write(row, 1, f"${data.get('npv_impact', 0):,.0f}")
        row += 2
        
        # Key Drivers
        if 'key_drivers' in data:
            sheet.write(row, 0, 'Key Risk Drivers', header_format)
            sheet.write(row, 1, '', header_format)
            row += 1
            for driver in data['key_drivers']:
                sheet.write(row, 0, driver)
                row += 1
        
        workbook.close()
        buffer.seek(0)
        return buffer
    
    def generate_scenario_analysis_excel(self, data: Dict) -> io.BytesIO:
        """Generate Excel report for Scenario Analysis."""
        buffer = io.BytesIO()
        workbook = xlsxwriter.Workbook(buffer, {'in_memory': True})
        
        # Formats
        header_format = workbook.add_format({
            'bold': True, 'bg_color': '#7c3aed', 'font_color': 'white',
            'align': 'center', 'border': 1
        })
        currency_format = workbook.add_format({'border': 1, 'num_format': '$#,##0'})
        percent_format = workbook.add_format({'border': 1, 'num_format': '+0.0%;-0.0%'})
        title_format = workbook.add_format({
            'bold': True, 'font_size': 16, 'font_color': '#7c3aed'
        })
        
        sheet = workbook.add_worksheet('Scenario Comparison')
        sheet.set_column('A:A', 25)
        sheet.set_column('B:E', 18)
        
        sheet.write('A1', 'Scenario Analysis Report', title_format)
        sheet.write('A2', f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}")
        
        if 'comparison_table' in data:
            row = 4
            headers = ['Scenario', 'Total Value', 'Change %', 'Risk Score', 'VaR 95%']
            for col, header in enumerate(headers):
                sheet.write(row, col, header, header_format)
            
            row += 1
            for scenario in data['comparison_table']:
                sheet.write(row, 0, scenario.get('scenario_name', 'N/A'))
                sheet.write(row, 1, scenario.get('total_value', 0), currency_format)
                sheet.write(row, 2, scenario.get('value_change_pct', 0) / 100, percent_format)
                sheet.write(row, 3, scenario.get('avg_risk_score', 0))
                sheet.write(row, 4, scenario.get('var_95', 0), currency_format)
                row += 1
            
            row += 1
            if 'best_scenario' in data:
                sheet.write(row, 0, 'Best Scenario:')
                sheet.write(row, 1, data['best_scenario'])
                row += 1
            if 'worst_scenario' in data:
                sheet.write(row, 0, 'Worst Scenario:')
                sheet.write(row, 1, data['worst_scenario'])
        
        workbook.close()
        buffer.seek(0)
        return buffer
    
    # ==================== Universal Export Method ====================
    
    def export(
        self, 
        module: str, 
        data: Dict, 
        format: str = 'pdf'
    ) -> io.BytesIO:
        """
        Universal export method that routes to appropriate generator.
        
        Args:
            module: One of 'portfolio_analytics', 'sustainability', 'stranded_assets', 
                   'scenario_analysis', 'nature_risk', 'valuation'
            data: The data to export
            format: 'pdf' or 'excel'
        
        Returns:
            BytesIO buffer with the generated file
        """
        generators = {
            'portfolio_analytics': (
                self.generate_portfolio_analytics_pdf,
                self.generate_portfolio_analytics_excel
            ),
            'sustainability': (
                self.generate_sustainability_pdf,
                self.generate_sustainability_excel
            ),
            'stranded_assets': (
                self.generate_stranded_assets_pdf,
                self.generate_stranded_assets_excel
            ),
            'scenario_analysis': (
                self.generate_scenario_analysis_pdf,
                self.generate_scenario_analysis_excel
            ),
        }
        
        if module not in generators:
            # Default to generic export
            if format == 'pdf':
                return self.generate_portfolio_analytics_pdf(data)
            return self.generate_portfolio_analytics_excel(data)
        
        pdf_gen, excel_gen = generators[module]
        
        if format == 'excel':
            return excel_gen(data)
        return pdf_gen(data)


# Singleton instance
export_service = ExportService()
