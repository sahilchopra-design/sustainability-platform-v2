"""
Sub-parameter analysis export — Excel/PDF/JSON.
"""

import os
import uuid
import json
from datetime import datetime, timezone
from typing import List, Dict, Any

REPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "reports_output")
os.makedirs(REPORTS_DIR, exist_ok=True)


def export_analysis_excel(analyses: List[Dict[str, Any]]) -> str:
    """Export analysis results to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    filename = f"analysis_export_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0F2137", end_color="0F2137", fill_type="solid")

    for i, analysis in enumerate(analyses):
        method = analysis.get("method", analysis.get("analysis_type", f"Analysis_{i}"))
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = method[:30]
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20

        # Header
        ws.append([f"Analysis: {method}"])
        ws.cell(1, 1).font = Font(bold=True, size=14, color="0F2137")
        ws.append([f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"])
        ws.append([])

        if "tornado_data" in analysis:
            ws.append(["Parameter", "Low Impact", "High Impact", "Swing", "Sensitivity %"])
            for c in range(1, 6):
                ws.cell(4, c).font = header_font
                ws.cell(4, c).fill = header_fill
            for td in analysis["tornado_data"]:
                ws.append([td["parameter"], td["low_impact"], td["high_impact"], td["swing"], td["sensitivity_score"]])

        elif "elasticities" in analysis:
            ws.append(["Parameter", "Elasticity", "Outcome Change %", "Interpretation"])
            for c in range(1, 5):
                ws.cell(4, c).font = header_font
                ws.cell(4, c).fill = header_fill
            for e in analysis["elasticities"]:
                ws.append([e["parameter"], e["elasticity"], e["outcome_change_pct"], e["interpretation"]])

        elif "correlations" in analysis:
            ws.append(["Parameter", "Correlation", "Direction", "Strength"])
            for c in range(1, 5):
                ws.cell(4, c).font = header_font
                ws.cell(4, c).fill = header_fill
            for cr in analysis["correlations"]:
                ws.append([cr["parameter"], cr["correlation"], cr["direction"], cr["strength"]])

        elif "coefficients" in analysis:
            ws.append(["Parameter", "Coefficient", "Weight %", "Direction"])
            for c in range(1, 5):
                ws.cell(4, c).font = header_font
                ws.cell(4, c).fill = header_fill
            for co in analysis["coefficients"]:
                ws.append([co["parameter"], co["coefficient"], co["weight_pct"], co["direction"]])

        elif "attributions" in analysis:
            ws.append(["Parameter", "Shapley Value", "Contribution %", "Direction"])
            for c in range(1, 5):
                ws.cell(4, c).font = header_font
                ws.cell(4, c).fill = header_fill
            for a in analysis["attributions"]:
                ws.append([a["parameter"], a["shapley_value"], a["contribution_pct"], a["direction"]])

        elif "differences" in analysis:
            ws.append(["Metric", "Baseline", "Modified", "Change %"])
            for c in range(1, 5):
                ws.cell(4, c).font = header_font
                ws.cell(4, c).fill = header_fill
            for k, d in analysis.get("differences", {}).items():
                ws.append([k, d.get("baseline"), d.get("modified"), d.get("pct_change")])

    wb.save(filepath)
    return filename


def export_analysis_pdf(analyses: List[Dict[str, Any]]) -> str:
    """Export analysis results to PDF."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.colors import HexColor, white
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable

    filename = f"analysis_export_{uuid.uuid4().hex[:8]}.pdf"
    filepath = os.path.join(REPORTS_DIR, filename)
    doc = SimpleDocTemplate(filepath, pagesize=A4, leftMargin=25*mm, rightMargin=25*mm)

    styles = getSampleStyleSheet()
    primary = HexColor("#0f2137")
    accent = HexColor("#06b6d4")

    h1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=16, textColor=primary, spaceAfter=8)
    h2 = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=12, textColor=accent, spaceAfter=6)
    body = styles["BodyText"]
    small = ParagraphStyle("Sm", parent=body, fontSize=8, textColor=HexColor("#6b7280"))

    elements = [
        Paragraph("A2 Intelligence — Sub-Parameter Analysis Report", h1),
        Paragraph(f"Generated: {datetime.now(timezone.utc).strftime('%B %d, %Y %H:%M UTC')}", small),
        HRFlowable(width="100%", thickness=2, color=primary),
        Spacer(1, 15),
    ]

    for analysis in analyses:
        method = analysis.get("method", analysis.get("analysis_type", "Analysis"))
        elements.append(Paragraph(f"{method.replace('_', ' ').title()} Analysis", h2))

        rows = []
        if "tornado_data" in analysis:
            rows = [["Parameter", "Low", "High", "Swing"]]
            rows += [[td["parameter"][:30], f"{td['low_impact']:.3f}", f"{td['high_impact']:.3f}", f"{td['swing']:.4f}"]
                     for td in analysis["tornado_data"][:10]]
        elif "elasticities" in analysis:
            rows = [["Parameter", "Elasticity", "Interpretation"]]
            rows += [[e["parameter"][:30], f"{e['elasticity']:.4f}", e["interpretation"][:50]]
                     for e in analysis["elasticities"][:10]]
        elif "coefficients" in analysis:
            rows = [["Parameter", "Coefficient", "Weight %"]]
            rows += [[c["parameter"][:30], f"{c['coefficient']:.6f}", f"{c['weight_pct']:.1f}%"]
                     for c in analysis["coefficients"][:10]]

        if rows:
            t = Table(rows, colWidths=[160, 80, 80, 120])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), primary),
                ("TEXTCOLOR", (0, 0), (-1, 0), white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#d1d5db")),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            elements.append(t)

        elements.append(Spacer(1, 15))

    doc.build(elements)
    return filename


def export_analysis_json(analyses: List[Dict[str, Any]]) -> str:
    """Export raw analysis data as JSON."""
    filename = f"analysis_export_{uuid.uuid4().hex[:8]}.json"
    filepath = os.path.join(REPORTS_DIR, filename)
    with open(filepath, "w") as f:
        json.dump({"analyses": analyses, "exported_at": datetime.now(timezone.utc).isoformat()}, f, indent=2, default=str)
    return filename
