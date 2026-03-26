"""
Report Generator — creates professional PDF and Excel reports for impact analysis.

Narrative sections use `narrative_templates.render_section()` so that all
boilerplate prose is maintained in a single place (narrative_templates.py).
"""

import os
import uuid
from datetime import datetime, timezone
from typing import Dict, Any, List

from services.narrative_templates import render_section

REPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "reports_output")
os.makedirs(REPORTS_DIR, exist_ok=True)


def _build_narrative_data(impact_data: dict, portfolio_data: dict, scenario_data: dict) -> dict:
    """Build the template data dict from raw report inputs."""
    horizons = impact_data.get("horizons", [])
    sc = scenario_data or {}

    # Find worst / best scenarios from horizons list
    sorted_h = sorted(horizons, key=lambda h: h.get("expected_loss", 0), reverse=True)
    worst = sorted_h[0] if sorted_h else {}
    best = sorted_h[-1] if sorted_h else {}

    sector_bd = portfolio_data.get("sectorBreakdown") or portfolio_data.get("sector_breakdown") or {}
    top_sector = max(sector_bd, key=sector_bd.get, default="N/A") if sector_bd else "N/A"
    top_sector_val = sector_bd.get(top_sector, 0)
    total_exp = portfolio_data.get("total_exposure", 1) or 1

    return {
        "portfolio_name": portfolio_data.get("name", "Portfolio"),
        "num_assets": portfolio_data.get("num_assets", 0),
        "currency": "$",
        "total_exposure": portfolio_data.get("total_exposure", 0),
        "num_scenarios": len(set(h.get("scenario_name", "") for h in horizons)),
        "horizons": ", ".join(str(h.get("horizon", "")) for h in horizons),
        "worst_scenario": impact_data.get("scenario_name", "Delayed Transition"),
        "worst_el": worst.get("expected_loss", 0),
        "worst_horizon": worst.get("horizon", 2050),
        "worst_el_pct": worst.get("expected_loss_pct", 0) * 100,
        "worst_pd_change": worst.get("avg_pd_change_pct", 0),
        "worst_var95": worst.get("var_95", 0),
        "best_scenario": impact_data.get("best_scenario_name", "Net Zero 2050"),
        "best_el": best.get("expected_loss", 0),
        "best_horizon": best.get("horizon", 2050),
        "best_el_pct": best.get("expected_loss_pct", 0) * 100,
        "top_sector": top_sector,
        "top_sector_pct": (top_sector_val / total_exp * 100) if total_exp else 0,
        "top_country": portfolio_data.get("top_country", "Germany"),
        "top_country_pct": portfolio_data.get("top_country_pct", 30),
        "top_country_hazard": portfolio_data.get("top_country_hazard", "riverine flooding"),
        "entity_name": portfolio_data.get("entity_name", portfolio_data.get("name", "the entity")),
        "reporting_year": datetime.now(timezone.utc).year,
    }


def generate_pdf_report(impact_data: dict, portfolio_data: dict, scenario_data: dict) -> str:
    """Generate a professional PDF report. Returns the filename."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import inch, mm
    from reportlab.lib.colors import HexColor, white, black
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        PageBreak, HRFlowable,
    )
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

    filename = f"impact_report_{uuid.uuid4().hex[:8]}.pdf"
    filepath = os.path.join(REPORTS_DIR, filename)

    doc = SimpleDocTemplate(filepath, pagesize=A4,
                            leftMargin=25*mm, rightMargin=25*mm,
                            topMargin=20*mm, bottomMargin=20*mm)

    styles = getSampleStyleSheet()
    primary = HexColor("#1e3a5f")
    accent = HexColor("#0d9488")
    light_bg = HexColor("#f0f4f8")
    red = HexColor("#dc2626")

    title_style = ParagraphStyle("Title", parent=styles["Title"], fontSize=22,
                                  textColor=primary, spaceAfter=6)
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=16,
                         textColor=primary, spaceBefore=18, spaceAfter=8)
    h2 = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=13,
                         textColor=accent, spaceBefore=14, spaceAfter=6)
    body = styles["BodyText"]
    small = ParagraphStyle("Small", parent=body, fontSize=8, textColor=HexColor("#6b7280"))

    elements = []

    # ---- Cover Page ----
    elements.append(Spacer(1, 60))
    elements.append(Paragraph("Climate Risk Impact Report", title_style))
    elements.append(Spacer(1, 12))
    elements.append(HRFlowable(width="100%", thickness=2, color=primary))
    elements.append(Spacer(1, 20))

    cover_data = [
        ["Portfolio:", portfolio_data.get("name", "—")],
        ["Assets:", str(portfolio_data.get("num_assets", "—"))],
        ["Total Exposure:", f"${portfolio_data.get('total_exposure', 0):,.0f}"],
        ["Scenario:", impact_data.get("scenario_name", "—")],
        ["Scenario Type:", impact_data.get("engine_scenario", "—")],
        ["Generated:", datetime.now(timezone.utc).strftime("%B %d, %Y %H:%M UTC")],
    ]
    cover_table = Table(cover_data, colWidths=[120, 340])
    cover_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("TEXTCOLOR", (0, 0), (0, -1), primary),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elements.append(cover_table)
    elements.append(PageBreak())

    # ---- Executive Summary ----
    elements.append(Paragraph("1. Executive Summary", h1))

    horizons = impact_data.get("horizons", [])
    if horizons:
        last = horizons[-1]
        elements.append(Paragraph(
            f"Under the <b>{impact_data.get('scenario_name', '')}</b> scenario "
            f"(mapped to <b>{impact_data.get('engine_scenario', '')}</b>), "
            f"the portfolio's expected loss reaches <b>${last.get('expected_loss', 0):,.0f}</b> "
            f"by {last.get('horizon', 2050)}, representing "
            f"<b>{last.get('expected_loss_pct', 0) * 100:.1f}%</b> of total exposure. "
            f"Value at Risk (95%) is <b>${last.get('var_95', 0):,.0f}</b>.",
            body
        ))
    elements.append(Spacer(1, 8))

    # Key metrics table
    if horizons:
        header = ["Metric"] + [str(h["horizon"]) for h in horizons]
        el_row = ["Expected Loss ($)"] + [f"${h.get('expected_loss', 0):,.0f}" for h in horizons]
        el_pct = ["EL % of Exposure"] + [f"{h.get('expected_loss_pct', 0) * 100:.1f}%" for h in horizons]
        var_row = ["VaR 95% ($)"] + [f"${h.get('var_95', 0):,.0f}" for h in horizons]
        pd_row = ["Avg PD Change (%)"] + [f"{h.get('avg_pd_change_pct', 0):.1f}%" for h in horizons]

        data = [header, el_row, el_pct, var_row, pd_row]
        col_w = [140] + [100] * len(horizons)
        t = Table(data, colWidths=col_w)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), primary),
            ("TEXTCOLOR", (0, 0), (-1, 0), white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#d1d5db")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, light_bg]),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(t)

    # ---- Narrative Summary (template-driven) ----
    try:
        narrative_data = _build_narrative_data(impact_data, portfolio_data, scenario_data)
        narrative_text = render_section("CLIMATE_RISK_IMPACT", narrative_data)
        elements.append(Paragraph("1a. Analytical Narrative", h2))
        for line in narrative_text.strip().split("\n"):
            if line.startswith("EXECUTIVE SUMMARY"):
                continue  # already have cover heading
            if line.strip():
                elements.append(Paragraph(line.strip(), body))
                elements.append(Spacer(1, 3))
    except Exception:
        pass  # narrative is best-effort — never block report generation

    # ---- Portfolio Overview ----
    elements.append(Paragraph("2. Portfolio Overview", h1))
    elements.append(Paragraph(
        f"The portfolio <b>{portfolio_data.get('name', '')}</b> contains "
        f"<b>{portfolio_data.get('num_assets', 0)}</b> assets with a total exposure of "
        f"<b>${portfolio_data.get('total_exposure', 0):,.0f}</b>.",
        body
    ))

    # ---- Scenario Description ----
    elements.append(Paragraph("3. Scenario Description", h1))
    sc = scenario_data or {}
    elements.append(Paragraph(
        f"<b>{sc.get('name', impact_data.get('scenario_name', ''))}</b>",
        body
    ))
    if sc.get("description"):
        elements.append(Paragraph(sc["description"], body))

    sc_details = []
    if sc.get("source_name"):
        sc_details.append(["Source:", sc["source_name"]])
    if sc.get("category"):
        sc_details.append(["Category:", sc["category"]])
    if sc.get("temperature_target"):
        sc_details.append(["Temperature Target:", f"{sc['temperature_target']}°C"])
    if sc.get("carbon_neutral_year"):
        sc_details.append(["Net Zero Year:", str(sc["carbon_neutral_year"])])
    if sc.get("model"):
        sc_details.append(["Model:", sc["model"]])

    if sc_details:
        st = Table(sc_details, colWidths=[120, 340])
        st.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(st)

    # Multipliers
    mults = impact_data.get("multipliers", {})
    if any(v is not None for v in mults.values()):
        elements.append(Paragraph("Scenario Multipliers", h2))
        mult_rows = [["Parameter", "Value"]]
        if mults.get("carbon_price_2030") is not None:
            mult_rows.append(["Carbon Price 2030", f"${mults['carbon_price_2030']:.1f}/tCO2"])
        if mults.get("carbon_price_2050") is not None:
            mult_rows.append(["Carbon Price 2050", f"${mults['carbon_price_2050']:.1f}/tCO2"])
        if mults.get("emissions_change_pct") is not None:
            mult_rows.append(["Emissions Change", f"{mults['emissions_change_pct']:+.1f}%"])
        if mults.get("temperature_2050") is not None:
            mult_rows.append(["Temperature 2050", f"{mults['temperature_2050']:.2f}°C"])

        mt = Table(mult_rows, colWidths=[200, 260])
        mt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), accent),
            ("TEXTCOLOR", (0, 0), (-1, 0), white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#d1d5db")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, light_bg]),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(mt)

    # ---- Impact Analysis Detail ----
    elements.append(Paragraph("4. Impact Analysis by Horizon", h1))
    for h in horizons:
        elements.append(Paragraph(f"Horizon: {h['horizon']}", h2))
        detail_data = [
            ["Expected Loss", f"${h.get('expected_loss', 0):,.0f}"],
            ["EL % of Exposure", f"{h.get('expected_loss_pct', 0) * 100:.2f}%"],
            ["VaR 95%", f"${h.get('var_95', 0):,.0f}"],
            ["VaR 99%", f"${h.get('var_99', 0):,.0f}"],
            ["Weighted Avg PD", f"{h.get('weighted_avg_pd', 0) * 100:.4f}%"],
            ["Avg PD Change", f"{h.get('avg_pd_change_pct', 0):.1f}%"],
            ["Total Exposure", f"${h.get('total_exposure', 0):,.0f}"],
            ["Risk-Adjusted Return", f"{h.get('risk_adjusted_return', 0) * 100:.2f}%"],
        ]
        dt = Table(detail_data, colWidths=[200, 260])
        dt.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#e5e7eb")),
            ("ROWBACKGROUNDS", (0, 0), (-1, -1), [white, light_bg]),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(dt)
        elements.append(Spacer(1, 6))

        # Rating migrations
        rm = h.get("rating_migrations", {})
        if rm:
            elements.append(Paragraph(
                f"Rating Migrations: ↑{rm.get('upgrades', 0)} upgrades, "
                f"→{rm.get('stable', 0)} stable, ↓{rm.get('downgrades', 0)} downgrades",
                small
            ))

    # ---- Footer ----
    elements.append(Spacer(1, 30))
    elements.append(HRFlowable(width="100%", thickness=1, color=HexColor("#d1d5db")))
    elements.append(Paragraph(
        f"Generated by Climate Risk Intelligence Platform — {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
        small
    ))

    doc.build(elements)
    return filename


def generate_excel_report(impact_data: dict, portfolio_data: dict, scenario_data: dict) -> str:
    """Generate an Excel report. Returns the filename."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    filename = f"impact_report_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    accent_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'), right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'), bottom=Side(style='thin', color='D1D5DB'),
    )

    # --- Summary Sheet ---
    ws = wb.active
    ws.title = "Executive Summary"
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30

    ws.append(["Climate Risk Impact Report"])
    ws['A1'].font = Font(bold=True, size=16, color="1E3A5F")
    ws.append([])
    ws.append(["Portfolio", portfolio_data.get("name", "")])
    ws.append(["Assets", portfolio_data.get("num_assets", 0)])
    ws.append(["Total Exposure", portfolio_data.get("total_exposure", 0)])
    ws.append(["Scenario", impact_data.get("scenario_name", "")])
    ws.append(["Engine Type", impact_data.get("engine_scenario", "")])
    ws.append(["Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")])
    for r in range(3, 9):
        ws.cell(r, 1).font = Font(bold=True)

    # --- Impact Results Sheet ---
    ws2 = wb.create_sheet("Impact Results")
    ws2.column_dimensions['A'].width = 22
    horizons = impact_data.get("horizons", [])

    headers = ["Metric"] + [str(h["horizon"]) for h in horizons]
    ws2.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws2.cell(1, c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    metrics = [
        ("Expected Loss ($)", "expected_loss"),
        ("EL % of Exposure", "expected_loss_pct"),
        ("VaR 95% ($)", "var_95"),
        ("VaR 99% ($)", "var_99"),
        ("Weighted Avg PD", "weighted_avg_pd"),
        ("Avg PD Change %", "avg_pd_change_pct"),
        ("Total Exposure ($)", "total_exposure"),
        ("Risk-Adjusted Return", "risk_adjusted_return"),
    ]
    for label, key in metrics:
        row = [label]
        for h in horizons:
            val = h.get(key, 0)
            if "pct" in key.lower() or "return" in key.lower():
                row.append(f"{val * 100:.2f}%" if "expected_loss_pct" in key or "return" in key else f"{val:.1f}%")
            else:
                row.append(val)
        ws2.append(row)

    for r in range(2, ws2.max_row + 1):
        ws2.cell(r, 1).font = Font(bold=True)
        for c in range(1, ws2.max_column + 1):
            ws2.cell(r, c).border = thin_border

    # --- Multipliers Sheet ---
    mults = impact_data.get("multipliers", {})
    if any(v is not None for v in mults.values()):
        ws3 = wb.create_sheet("Scenario Multipliers")
        ws3.column_dimensions['A'].width = 25
        ws3.column_dimensions['B'].width = 20
        ws3.append(["Parameter", "Value"])
        ws3.cell(1, 1).font = header_font
        ws3.cell(1, 1).fill = accent_fill
        ws3.cell(1, 2).font = header_font
        ws3.cell(1, 2).fill = accent_fill

        if mults.get("carbon_price_2030") is not None:
            ws3.append(["Carbon Price 2030", mults["carbon_price_2030"]])
        if mults.get("carbon_price_2050") is not None:
            ws3.append(["Carbon Price 2050", mults["carbon_price_2050"]])
        if mults.get("emissions_change_pct") is not None:
            ws3.append(["Emissions Change %", mults["emissions_change_pct"]])
        if mults.get("temperature_2050") is not None:
            ws3.append(["Temperature 2050 °C", mults["temperature_2050"]])

    # --- Scenario Info Sheet ---
    sc = scenario_data or {}
    ws4 = wb.create_sheet("Scenario Details")
    ws4.column_dimensions['A'].width = 20
    ws4.column_dimensions['B'].width = 50
    info = [
        ("Name", sc.get("name", impact_data.get("scenario_name", ""))),
        ("Source", sc.get("source_name", "")),
        ("Category", sc.get("category", "")),
        ("Temperature Target", f"{sc.get('temperature_target', '')}°C" if sc.get("temperature_target") else ""),
        ("Carbon Neutral Year", sc.get("carbon_neutral_year", "")),
        ("Model", sc.get("model", "")),
        ("Description", sc.get("description", "")),
    ]
    ws4.append(["Field", "Value"])
    ws4.cell(1, 1).font = header_font
    ws4.cell(1, 1).fill = header_fill
    ws4.cell(1, 2).font = header_font
    ws4.cell(1, 2).fill = header_fill
    for k, v in info:
        ws4.append([k, str(v) if v else ""])

    wb.save(filepath)
    return filename


def generate_csv_report(impact_data: dict, portfolio_data: dict) -> str:
    """Generate a CSV report of impact results."""
    import csv
    import io

    filename = f"impact_report_{uuid.uuid4().hex[:8]}.csv"
    filepath = os.path.join(REPORTS_DIR, filename)

    with open(filepath, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Climate Risk Impact Report"])
        writer.writerow(["Portfolio", portfolio_data.get("name", "")])
        writer.writerow(["Scenario", impact_data.get("scenario_name", "")])
        writer.writerow(["Engine Type", impact_data.get("engine_scenario", "")])
        writer.writerow([])
        writer.writerow(["Horizon", "Expected Loss", "EL %", "VaR 95%", "VaR 99%",
                         "Avg PD Change %", "Weighted Avg PD", "Total Exposure"])
        for h in impact_data.get("horizons", []):
            writer.writerow([
                h.get("horizon"), h.get("expected_loss"), h.get("expected_loss_pct"),
                h.get("var_95"), h.get("var_99"), h.get("avg_pd_change_pct"),
                h.get("weighted_avg_pd"), h.get("total_exposure"),
            ])

    return filename
